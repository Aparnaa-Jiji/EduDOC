# =========================================================
# STUDENT + BATCH ADMIN VIEWS – EduDOC
# Consolidated • Clean • Stable • Production Safe
# =========================================================
from teacher.services.submission_gate import get_submission_gate
from datetime import timedelta
import csv
import io
import random
import string
from django.urls import reverse
from django.conf import settings
from teacher.services.evaluator import DocumentEvaluator

from openpyxl import load_workbook

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.core.mail import send_mail
from django.http import HttpResponse
from django.contrib.auth import get_user_model


from teacher.models import Classroom, Submission

from .models import (
    Student,
    ClassroomMembership,
    UnlockRequest,
    Batch
)

from .forms import (
    JoinClassroomForm,
    SubmissionForm,
    ManualStudentCreateForm,
    BatchForm
)

User = get_user_model()


# =========================================================
# EMAIL HELPER
# =========================================================

def send_student_credentials(request, email, username, password):

    login_url = request.build_absolute_uri(
        reverse("accounts:login")
    )

    send_mail(
        "EduDOC Student Account Credentials",
        f"""
Welcome to EduDOC

Username: {username}
Password: {password}

Login: {login_url}

Please log in and change your password immediately.
""",
        settings.DEFAULT_FROM_EMAIL,
        [email],
        fail_silently=True
    )


# =========================================================
# ROLE DECORATOR
# =========================================================
def student_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated or request.user.role != User.Role.STUDENT:
            return redirect("accounts:login")
        return view_func(request, *args, **kwargs)
    return wrapper

from django.db.models import Max
# =========================================================
# ================= STUDENT MODULE ========================
# =========================================================
@login_required
@student_required
def student_dashboard(request):

    student = get_object_or_404(Student, user=request.user)
    now = timezone.now()
    next_48_hours = now + timedelta(hours=48)

    memberships = ClassroomMembership.objects.filter(
        student=student
    ).select_related("classroom__teacher")

    total_joined = memberships.count()

    active_count = 0
    due_soon_count = 0
    evaluated_count = 0
    pending_count = 0
    total_score = 0
    score_count = 0

    urgent_rows = []

    for m in memberships:

        classroom = m.classroom

      

        # Get latest attempt per student
        latest_attempts = (
            Submission.objects
            .filter(classroom=classroom)
            .values("student")
            .annotate(max_attempt=Max("attempt_no"))
        )

        submissions = Submission.objects.filter(
            classroom=classroom,
            attempt_no__in=[item["max_attempt"] for item in latest_attempts]
        )


        latest = submissions.first()
        attempts_used = submissions.count()

        deadline_passed = classroom.deadline and classroom.deadline < now

        if classroom.deadline and classroom.deadline >= now:
            active_count += 1

        if classroom.deadline and now <= classroom.deadline <= next_48_hours:
            due_soon_count += 1

            urgent_rows.append({
                "classroom": classroom,
                "teacher": classroom.teacher.username,
                "deadline": classroom.deadline,
                "latest_submission": latest,
                "attempts_used": attempts_used,
                "attempts_allowed": classroom.attempts_allowed,
            })

        if latest and latest.status == "EVALUATED":
            evaluated_count += 1
            if latest.final_score:
                total_score += latest.final_score
                score_count += 1

        if not latest and classroom.deadline and classroom.deadline >= now:
            pending_count += 1

    avg_score = int(total_score / score_count) if score_count else 0

    return render(request, "students/student_dashboard.html", {
        "total_joined": total_joined,
        "active_count": active_count,
        "due_soon_count": due_soon_count,
        "evaluated_count": evaluated_count,
        "pending_count": pending_count,
        "avg_score": avg_score,
        "urgent_rows": urgent_rows,
    })

# ---------------------------------------------------------
# Join Classroom
# ---------------------------------------------------------
@login_required
@student_required
def join_classroom(request):

    student = get_object_or_404(Student, user=request.user)

    form = JoinClassroomForm(request.POST or None)

    if request.method == "POST" and form.is_valid():

        passkey = form.cleaned_data["passkey"]

        try:
            classroom = Classroom.objects.get(passkey=passkey)
            if classroom.batch != student.batch:
                messages.error(request, "You cannot join this classroom.")
                return redirect("student:student_dashboard")
            ClassroomMembership.objects.get_or_create(
                student=student,
                classroom=classroom
            )

            messages.success(request, "Joined successfully.")

        except Classroom.DoesNotExist:
            messages.error(request, "Invalid passkey.")

    return redirect("student:student_dashboard")


# ---------------------------------------------------------
# Classroom Workspace
# ---------------------------------------------------------

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from datetime import timedelta

from accounts.decorators import student_required
from student.models import Student, ClassroomMembership
from teacher.models import Classroom
from teacher.models import Submission



from django.shortcuts import render, get_object_or_404
from django.utils import timezone

from student.models import Student, ClassroomMembership
from teacher.models import Classroom, Submission
from teacher.services.submission_gate import get_submission_gate

from .forms import SubmissionForm


def classroom_workspace(request, classroom_id):

    # =====================================================
    # FETCH CORE OBJECTS
    # =====================================================
    student = get_object_or_404(Student, user=request.user)

    classroom = get_object_or_404(
        Classroom,
        id=classroom_id
    )

    membership = get_object_or_404(
        ClassroomMembership,
        student=student,
        classroom=classroom
    )

    now = timezone.now()

    # =====================================================
    # UNLOCK STATUS
    # =====================================================
    unlock_active = (
        membership.unlock_until and
        membership.unlock_until > now
    )

    # =====================================================
    # DEADLINE CHECK (UNLOCK OVERRIDE)
    # =====================================================
    deadline_passed = (
        classroom.deadline and now > classroom.deadline
    )

    if unlock_active:
        deadline_passed = False

    # =====================================================
    # CENTRALIZED SUBMISSION GATE ✅
    # =====================================================
    gate = get_submission_gate(
        student=student,
        classroom=classroom,
        unlock_active=unlock_active
    )

    can_submit = gate["allowed"]
    cooldown_remaining_seconds = gate["remaining_seconds"]

    attempts_allowed = gate["attempts_allowed"]
    attempts_used = gate["attempts_used"]

    cooldown_active = not gate["allowed"]

    attempts_remaining = max(
        0,
        attempts_allowed - attempts_used
    )

    # =====================================================
    # LOCK REASON
    # =====================================================
    submission_locked_reason = None

    if deadline_passed:
        can_submit = False
        submission_locked_reason = "Deadline passed"

    elif cooldown_active:
        submission_locked_reason = "Attempt cooldown active"

    # =====================================================
    # FETCH SUBMISSIONS (DISPLAY ONLY)
    # =====================================================
    submissions = list(
        Submission.objects.filter(
            student=student,
            classroom=classroom
        ).order_by("-submitted_at")
    )

    latest_submission = submissions[0] if submissions else None

    # =====================================================
    # PROGRESS BAR
    # =====================================================
    submission_progress = (
        int((attempts_used / attempts_allowed) * 100)
        if attempts_allowed else 0
    )

    # =====================================================
    # CONTEXT
    # =====================================================
    context = {
        "classroom": classroom,
        "membership": membership,
        "form": SubmissionForm(),

        "submissions": submissions,
        "latest_submission": latest_submission,

        "can_submit": can_submit,
        "submission_locked_reason": submission_locked_reason,

        "attempts_used": attempts_used,
        "attempts_allowed": attempts_allowed,
        "attempts_remaining": attempts_remaining,

        "cooldown_active": cooldown_active,
        "cooldown_remaining_seconds": cooldown_remaining_seconds,

        "deadline_passed": deadline_passed,
        "unlock_active": unlock_active,

        "submission_progress": submission_progress,
    }

    return render(
        request,
        "students/classroom_workspace.html",
        context
    )

# ---------------------------------------------------------
# Upload Submission
# ---------------------------------------------------------

from django.db import transaction
from django.db.models import Max


# =========================================================
# Upload Submission + Compliance Check (Phase-2)
# =========================================================



# =========================================================
# Upload Submission
# FINAL – Production Safe + Structured Evaluation
# =========================================================
@login_required
@student_required
def upload_submission(request, classroom_id):

    student = get_object_or_404(Student, user=request.user)
    classroom = get_object_or_404(Classroom, id=classroom_id)

    membership = get_object_or_404(
        ClassroomMembership,
        student=student,
        classroom=classroom
    )

    if request.method != "POST":
        return redirect("student:classroom_workspace",
                        classroom_id=classroom.id)

    now = timezone.now()

    # =====================================================
    # DEADLINE + UNLOCK STATUS (FIRST)
    # =====================================================
    deadline_passed = classroom.deadline and now > classroom.deadline

    unlock_active = (
        membership.unlock_until and
        membership.unlock_until > now
    )

    if deadline_passed and not unlock_active:
        
        messages.error(request, "Deadline passed. Request unlock.")
        return redirect(
            "student:classroom_workspace",
            classroom_id=classroom.id
        )

    # =====================================================
    # ATTEMPT GATE CHECK (CENTRALIZED)
    # =====================================================
    gate = get_submission_gate(
        student=student,
        classroom=classroom,
        unlock_active=unlock_active
    )

    if not gate["allowed"]:

        remaining = gate["remaining_seconds"]

        hours = remaining // 3600
        minutes = (remaining % 3600) // 60
        seconds = remaining % 60

        messages.error(
            request,
            f"Upload locked. Try again in "
            f"{hours:02d}:{minutes:02d}:{seconds:02d}"
        )

        return redirect(
            "student:classroom_workspace",
            classroom_id=classroom.id
        )

    # =====================================================
    # FORM VALIDATION
    # =====================================================
    form = SubmissionForm(request.POST, request.FILES)

    if not form.is_valid():
        for errors in form.errors.values():
            for error in errors:
                messages.error(request, error)

        return redirect(
            "student:classroom_workspace",
            classroom_id=classroom.id
        )

    # =====================================================
    # SAFE DATABASE TRANSACTION
    # =====================================================
    with transaction.atomic():

        submissions = Submission.objects.filter(
            student=student,
            classroom=classroom
        ).order_by("submitted_at")

        next_attempt = submissions.count() + 1

        submission = form.save(commit=False)
        submission.student = student
        submission.classroom = classroom
        submission.attempt_no = next_attempt
        submission.status = "PROCESSING"
        submission.save()

        # ---------- RULES ----------
        rules = classroom.rules_json or {}

        if not rules:
            submission.status = "ERROR"
            submission.save(update_fields=["status"])

            messages.error(request, "Evaluation rules not configured.")
            return redirect(
                "student:classroom_workspace",
                classroom_id=classroom.id
            )

        # ---------- EVALUATION ----------
        evaluator = DocumentEvaluator(
            file_path=submission.file.path,
            rules=rules,
            plagiarism_enabled=classroom.plagiarism_enabled
        )

        result = evaluator.evaluate()

        submission.breakdown_json = result
        submission.compliance_percent = result.get("compliance_percent", 0)
        submission.plagiarism_percent = result.get("plagiarism_percent", 0)
        submission.final_score = submission.compliance_percent
        submission.status = "EVALUATED"

        submission.save(update_fields=[
            "breakdown_json",
            "compliance_percent",
            "plagiarism_percent",
            "final_score",
            "status"
        ])

        # ---------- ANNOTATED FILE ----------
        if submission.file.path.endswith(".docx"):

            from django.core.files import File
            import os

            annotated_name = f"annotated_{submission.id}.docx"
            annotated_path = os.path.join(
                os.path.dirname(submission.file.path),
                annotated_name
            )

            evaluator.generate_annotated_file(annotated_path)

            with open(annotated_path, "rb") as f:
                submission.annotated_file.save(
                    annotated_name,
                    File(f),
                    save=True
                )

    messages.success(request, "Uploaded and analyzed successfully.")

    return redirect(
        "student:classroom_workspace",
        classroom_id=classroom.id
    )
# ---------------------------------------------------------
# Unlock Request
# ---------------------------------------------------------
@login_required
@student_required
def request_unlock(request, classroom_id):

    student = get_object_or_404(Student, user=request.user)
    classroom = get_object_or_404(Classroom, id=classroom_id)

    reason = request.POST.get("reason", "").strip()

    if not reason:
        messages.error(request, "Reason is required.")
        return redirect("student:classroom_workspace", classroom_id=classroom.id)

    UnlockRequest.objects.create(
        student=student,
        classroom=classroom,
        reason=reason
    )


    messages.success(request, "Request sent.")

    return redirect("student:classroom_workspace", classroom_id=classroom.id)


# =========================================================
# ================= ADMIN – BATCH MODULE ==================
# =========================================================

def admin_required(request):
    return request.user.role == User.Role.ADMIN








# =========================================================
# CSV/XLSX PARSER
# =========================================================
def parse_uploaded_students(file):

    filename = file.name.lower()

    if filename.endswith(".csv"):
        text = file.read().decode("utf-8")
        reader = csv.DictReader(io.StringIO(text))
        return list(reader)

    elif filename.endswith(".xlsx"):
        wb = load_workbook(file)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        headers = rows[0]
        data_rows = rows[1:]

        return [dict(zip(headers, row)) for row in data_rows]

    else:
        raise ValueError("Only CSV or XLSX supported")


# =========================================================
# BULK UPLOAD (FIXED)
# =========================================================
# ---------------------------------------------------------
# Bulk Upload Students (CSV / XLSX) + EMAIL (ADMIN)
# ---------------------------------------------------------
@login_required
def student_bulk_upload(request, id):

    if request.user.role != User.Role.ADMIN:
        return redirect("accounts:admin_dashboard")

    batch = get_object_or_404(Batch, id=id)

    file = request.FILES.get("file")

    if not file:
        messages.error(request, "No file uploaded.")
        return redirect("student:batch_detail", id=batch.id)

    try:
        rows = parse_uploaded_students(file)
    except Exception:
        messages.error(request, "Invalid file. Upload CSV or XLSX.")
        return redirect("student:batch_detail", id=batch.id)

    created = 0
    skipped = 0
    skipped_emails = []
    for row in rows:

        email = (row.get("email") or row.get("Email") or "").strip()
        name = (row.get("full_name") or row.get("name") or "").strip()

        if not email:
            continue

        # skip duplicates
        if User.objects.filter(email=email).exists():
            skipped += 1
            skipped_emails.append(email)
            continue

        base_username = email.split("@")[0]
        username = base_username
        counter = 1

        while User.objects.filter(username=username).exists():
            username = f"{base_username}{counter}"
            counter += 1

        password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))

        # create auth user
        full_name = name.strip()

        name_parts = full_name.split(" ", 1)
        first_name = name_parts[0] if name_parts else ""
        last_name = name_parts[1] if len(name_parts) > 1 else ""

        user = User.objects.create_user(
            username=username,
            email=email,
            password=password,
            role=User.Role.STUDENT,
            first_name=first_name,
            last_name=last_name
        )


        Student.objects.create(user=user,batch=batch)



        # 🔥 SEND EMAIL HERE (missing earlier)
        send_student_credentials(request, email, username, password)


        created += 1

    if skipped_emails:
        messages.warning(
            request,
            f"Upload complete — Created: {created}. Skipped {skipped} duplicate email(s): {', '.join(skipped_emails)}"
        )
    else:
        messages.success(
            request,
            f"Upload complete — Created: {created}. Credentials emailed."
        )


    return redirect("accounts:admin_batch_detail", id=batch.id)

# =========================================================
# EXPORT CSV
# =========================================================
@login_required
def batch_export_csv(request, id):

    batch = get_object_or_404(Batch, id=id)

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{batch.name}.csv"'

    writer = csv.writer(response)
    writer.writerow(["Email", "Username"])


    for s in batch.students.select_related("user"):
        writer.writerow([
            s.user.email,
            s.user.username
        ])

    return response
# ---------------------------------------------------------
# Report Detail
# ---------------------------------------------------------
def report_detail(request, submission_id):

    student = get_object_or_404(Student, user=request.user)

    submission = get_object_or_404(
        Submission,
        id=submission_id,
        student=student
    )

    breakdown = submission.breakdown_json or {}
    issues = breakdown.get("issues", [])

    # Add cm conversion for margin display
    for issue in issues:
        if issue.get("rule") == "Margins":

            for side in ["top", "bottom", "left", "right"]:

                # Expected
                inch_val = issue["expected"].get(side)
                if inch_val is not None:
                    issue["expected"][f"{side}_cm"] = round(inch_val * 2.54, 2)

                # Found
                found_val = issue["found"].get(side)
                if found_val is not None:
                    issue["found"][f"{side}_cm"] = round(found_val * 2.54, 2)

    return render(
        request,
        "students/report_detail.html",
        {
            "submission": submission,
        }
    )

# ---------------------------------------------------------
# Add Single Student (ADMIN)
# ---------------------------------------------------------
@login_required
def student_add(request, id):

    if request.user.role != User.Role.ADMIN:
        return redirect("accounts:admin_dashboard")

    batch = get_object_or_404(Batch, id=id)

    form = ManualStudentCreateForm(request.POST or None)

    if request.method == "POST" and form.is_valid():

        email = form.cleaned_data["email"]
      

        # prevent duplicates
        if User.objects.filter(email=email).exists():
            messages.error(request, "Email already exists.")
            return redirect("accounts:admin_batch_detail", id=batch.id)

        username = email.split("@")[0]
        password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))

        full_name = form.cleaned_data["full_name"].strip()

        name_parts = full_name.split(" ", 1)
        first_name = name_parts[0]
        last_name = name_parts[1] if len(name_parts) > 1 else ""

        user = User.objects.create_user(
            username=username,
            email=email,
            password=password,
            role=User.Role.STUDENT,
            first_name=first_name,
            last_name=last_name
        )


        # sequential register number (cleaner than random)
        Student.objects.create(
        user=user,
        batch=batch
    )

        send_student_credentials(request, email, username, password)


        messages.success(request, "Student created successfully.")

        return redirect("accounts:admin_batch_detail", id=batch.id)

    return render(
        request,
        "students/batches/student_add.html",
        {
            "form": form,
            "batch": batch
        }
    )

# ---------------------------------------------------------
# Delete Student (ADMIN)
# ---------------------------------------------------------
@login_required
def student_delete(request, id):

    if request.user.role != User.Role.ADMIN:
        return redirect("accounts:admin_dashboard")

    student = get_object_or_404(Student, id=id)

    batch_id = student.batch.id

    # delete user → cascades Student profile automatically
    student.user.delete()

    messages.success(request, "Student deleted successfully.")

    return redirect("accounts:admin_batch_detail", id=batch_id)

from django.utils import timezone

from django.utils import timezone
from django.db.models import Q
from datetime import timedelta

@login_required
@student_required
def classroom_list(request):

    student = get_object_or_404(Student, user=request.user)

    memberships = ClassroomMembership.objects.filter(
        student=student
    ).select_related("classroom")

    joined_ids = memberships.values_list("classroom_id", flat=True)

    now = timezone.now()
    # =====================================================
    # DEADLINE WARNING (< 24 HOURS)
    # =====================================================

    deadline_warning_time = now + timedelta(hours=24)

    urgent_classrooms = memberships.filter(
        classroom__deadline__gte=now,
        classroom__deadline__lte=deadline_warning_time
    )
    # ✅ ONLY ACTIVE CLASSROOMS
    active_memberships = memberships.filter(
        Q(classroom__deadline__gte=now) |
        Q(classroom__deadline__isnull=True)
    )

    # Available classrooms
    other_classrooms = Classroom.objects.filter(
        batch=student.batch
    ).exclude(id__in=joined_ids)

    context = {
        "memberships": memberships,
        "active_memberships": active_memberships,
        "other_classrooms": other_classrooms,
        "now": now,
        "urgent_classrooms": urgent_classrooms,
    }

    return render(
        request,
        "students/classroom_list.html",
        context
    )
@login_required
@student_required
def unjoin_classroom(request):

    if request.method == "POST":

        classroom_id = request.POST.get("classroom_id")

        student = get_object_or_404(Student, user=request.user)

        membership = get_object_or_404(
            ClassroomMembership,
            student=student,
            classroom_id=classroom_id
        )

        membership.delete()

        messages.success(request, "You have left the classroom.")

    return redirect("student:classroom_list")

@login_required
@student_required
def download_rules(request, classroom_id):

    student = get_object_or_404(Student, user=request.user)
    classroom = get_object_or_404(
        Classroom,
        id=classroom_id,
        batch=student.batch  # 🔒 batch scoped security
    )

    if not classroom.rules_json:
        messages.error(request, "No rules configured.")
        return redirect("student:classroom_workspace", classroom_id=classroom.id)

    # Reuse teacher PDF generator logic if needed
    from teacher.views import export_rules_pdf

    return export_rules_pdf(request, classroom_id)

from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Table, TableStyle
from reportlab.lib.units import inch
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics



def download_report_pdf(request, submission_id):
    submission = get_object_or_404(
        Submission,
        id=submission_id,
        student__user=request.user
    )

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="report_attempt_{submission.attempt_no}.pdf"'
    )

    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []

    styles = getSampleStyleSheet()

    # Title
    elements.append(Paragraph(
        "Academic Compliance Report",
        styles['Heading1']
    ))
    elements.append(Spacer(1, 0.3 * inch))

    # Summary Table
    summary_data = [
        ["Compliance", f"{submission.compliance_percent}%"],
        ["Plagiarism", f"{submission.plagiarism_percent}%"],
        ["Final Score", f"{submission.final_score}%"],
    ]

    summary_table = Table(summary_data, colWidths=[2 * inch, 2 * inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.whitesmoke),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
    ]))

    elements.append(summary_table)
    elements.append(Spacer(1, 0.5 * inch))

    # Rule Table
    rule_data = [["Rule", "Status", "Expected", "Detected"]]

    for issue in submission.breakdown_json.get("issues", []):
        rule_data.append([
            issue.get("rule", ""),
            issue.get("status", ""),
            str(issue.get("expected", "")),
            str(issue.get("found", "")),
        ])

    rule_table = Table(rule_data, repeatRows=1)
    rule_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
        ('FONTSIZE', (0,0), (-1,-1), 8),
    ]))

    elements.append(rule_table)

    doc.build(elements)

    return response

from django.http import FileResponse, HttpResponse
from django.shortcuts import get_object_or_404
from teacher.models import Submission
from django.conf import settings
import os


def download_annotated(request, submission_id):

    submission = get_object_or_404(Submission, id=submission_id)

    if not submission.annotated_file:
        return HttpResponse("Annotated file not available", status=404)

    file_path = os.path.join(settings.MEDIA_ROOT, submission.annotated_file.name)

    if not os.path.exists(file_path):
        return HttpResponse("File not found on server", status=404)

    return FileResponse(
        open(file_path, "rb"),
        as_attachment=True,
        filename=os.path.basename(file_path)
    )